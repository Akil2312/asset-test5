import streamlit as st
import openpyxl
import configparser
import bcrypt

# ----------------- Read config -----------------
cfg = configparser.ConfigParser()
cfg.read("config.txt")
EXCEL_FILE = cfg.get("DEFAULT", "EXCEL_FILE")

# ----------------- Utility functions -----------------
def load_users():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Users"]
    users = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        users.append(
            {"username": row[0], "password_hash": row[1], "role": row[2]}
        )
    return users


def load_assets():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Assets"]
    assets = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        assets.append(
            {
                "id": row[0],
                "username": row[1],
                "name": row[2],
                "status": row[3],
            }
        )
    return assets


def update_asset_status(asset_id, status):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Assets"]
    for row in ws.iter_rows(min_row=2):
        if str(row[0].value) == str(asset_id):
            row[3].value = status
    wb.save(EXCEL_FILE)


# ----------------- App Flow -----------------
def main():
    st.title("Asset Management System")

    # Sidebar for navigation
    menu = ["Login", "Approver", "IT User", "End User", "Logout"]
    choice = st.sidebar.selectbox("Menu", menu)

    if "logged_in" not in st.session_state:
        st.session_state["logged_in"] = False
        st.session_state["role"] = None
        st.session_state["user"] = None

    if choice == "Login":
        if not st.session_state["logged_in"]:
            username = st.text_input("Username")
            password = st.text_input("Password", type="password")

            if st.button("Login"):
                user = next(
                    (
                        u
                        for u in load_users()
                        if u["username"] == username
                    ),
                    None,
                )
                if user and bcrypt.checkpw(
                    password.encode(), user["password_hash"].encode()
                ):
                    st.session_state["logged_in"] = True
                    st.session_state["role"] = user["role"].lower()
                    st.session_state["user"] = user
                    st.success(
                        f"Welcome {st.session_state['user']['username']}!"
                    )
                else:
                    st.error("Invalid credentials")
        else:
            st.write("You are already logged in.")

    elif choice == "Logout":
        st.session_state["logged_in"] = False
        st.session_state["role"] = None
        st.session_state["user"] = None
        st.success("Logged out successfully.")

    # Approver Page
    elif choice == "Approver" and st.session_state["logged_in"]:
        if st.session_state["role"] == "approver":
            st.subheader("Assets Pending Approval")
            assets = [
                a for a in load_assets() if a["status"] == "Pending Approval"
            ]
            for asset in assets:
                st.write(
                    f"ID: {asset['id']}, Name: {asset['name']}, Status: {asset['status']}"
                )
                if st.button(
                    f"Approve-{asset['id']}",
                    key=f"approve-{asset['id']}",
                ):
                    update_asset_status(asset["id"], "Approved")
                    st.success("Asset Approved.")

                if st.button(
                    f"Reject-{asset['id']}", key=f"reject-{asset['id']}"
                ):
                    update_asset_status(asset["id"], "Rejected")
                    st.success("Asset Rejected.")
        else:
            st.error("Unauthorized access.")

    # IT User Page
    elif choice == "IT User" and st.session_state["logged_in"]:
        if st.session_state["role"] == "ituser":
            st.subheader("IT User Panel")
            username_search = st.text_input("Search by Username")

            if username_search:
                assets = [
                    a
                    for a in load_assets()
                    if a["username"] == username_search
                ]
                for asset in assets:
                    status = st.selectbox(
                        f"Set Status for {asset['name']}",
                        options=["Available", "In Use", "Pending Approval"],
                        index=0,
                        key=f"status-{asset['id']}",
                    )
                    if st.button(
                        f"Update-{asset['id']}",
                        key=f"update-{asset['id']}",
                    ):
                        update_asset_status(asset["id"], status)
                        st.success(f"Status updated for Asset ID: {asset['id']}")
        else:
            st.error("Unauthorized access.")

    # End User Page
    elif choice == "End User" and st.session_state["logged_in"]:
        if st.session_state["role"] == "enduser":
            st.subheader("Your Assets")
            username = st.session_state["user"]["username"]
            assets = [
                a for a in load_assets() if a["username"] == username
            ]
            for asset in assets:
                st.write(
                    f"ID: {asset['id']}, Name: {asset['name']}, Status: {asset['status']}"
                )
        else:
            st.error("Unauthorized access.")


if __name__ == "__main__":
    main()
